# -*- coding: utf-8 -*-
"""
topic: 커리큘럼 기반 토픽 유사도 인덱스

개선 사항(불용어 처리):
- 토크나이저(_tokenize)에서 stopword/숫자/날짜 관련 토큰을 필터링
  * 예: "휴강", "휴강일", "추석", "특강", "크리스마스", "혹은", "여러", "주차", "요일",
        "오전", "오후", "OT/오티", "수료식", "공지", "설문", 요일/월 이름, 순수 숫자("01","1"),
        연/월/일 형태("1월","12월","2025" → 하이픈/슬래시 포함시 분해되어도 숫자는 버림)
- timetable 전용 파서 유지(왼쪽 표만 사용) + 일반 파서 폴백

의존:
- openpyxl, numpy, rank-bm25
"""
from __future__ import annotations

import math
import re
from typing import List, Tuple, Dict, Optional, Any
from datetime import datetime

from openpyxl import load_workbook

import numpy as np
from rank_bm25 import BM25Okapi

# --------------------------
# Stopword & 패턴
# --------------------------
# 한/영 공통 불용어(시간표/행정/연휴/형식)
STOP_KO = {
    "휴강",
    "휴강일",
    "연휴",
    "공휴일",
    "추석",
    "설날",
    "크리스마스",
    "특강",
    "특강1",
    "특강2",
    "특강3",
    "수료식",
    "오티",
    "ot",
    "오리엔테이션",
    "주차",
    "일자",
    "요일",
    "구분",
    "오전",
    "오후",
    "점심",
    "저녁",
    "공지",
    "공지사항",
    "설문",
    "마무리",
    "정리",
    "여기까지",
    "소개",
    "온라인",
    "오프라인",
    "줌",
    "zoom",
    "마이크",
    "카메라",
    "화면",
    "공유",
    "혹은",
    "여러",
    "등",
    "및",
    "또는",
    "활용",
    "향후",
    "현업",
    "확장",
    "환경",
    "활용법",
    "활용한",
    "흐름",
    "학습",
    "학습법",
    "활용하여",
    "프로젝트",
    "프로젝트1",
    "프로젝트2",
    "환경에서의",
    "서비스",
    "서비스의",
    "수준의",
    "사용한",
    "빈출",
    "사항",
    "기타",
    "광복절",
    "노하우",
    "등을",
    "습득",
    "적용한",
    "학습을",
    "트렌드",
    "주요",
    "지향",
    "과정지식을",
    "관련",
    "월별",
    "위한",
    "우한",
    "코칭",
}
STOP_EN = {
    "am",
    "pm",
    "mon",
    "tue",
    "wed",
    "thu",
    "fri",
    "sat",
    "sun",
    "online",
    "offline",
    "zoom",
}
STOP_DAYS = {"월", "화", "수", "목", "금", "토", "일"}
STOP_MONTH_SUFFIX = {"월", "일", "년"}

# 기술 키워드 화이트리스트(숫자 포함/짧은 토큰 유지)
WHITELIST = {
    "ai",
    "ml",
    "aws",
    "gcp",
    "azure",
    "devops",
    "cicd",
    "ci",
    "cd",
    "api",
    "rest",
    "s3",
    "ec2",
    "eks",
    "vpc",
    "dns",
    "http",
    "tcp",
    "udp",
    "git",
    "github",
    "docker",
    "kubernetes",
    "k8s",
    "java",
    "python",
    "linux",
    "bash",
    "shell",
    "sql",
    "nosql",
    "rds",
    "iam",
    "sso",
    "oauth",
    "cloudfront",
    "cdn",
    "lambda",
    "terraform",
    "ansible",
    "nginx",
    "redis",
    "mongodb",
    "postgres",
    "mysql",
    "c",
    "c++",
    "c#",
    ".net",
}

_TOKEN_RE = re.compile(r"[A-Za-z가-힣0-9]+")

_NUMERIC_FULL = re.compile(r"^\d{1,4}$")  # 1~4자리 숫자만 (01, 1, 2025 등)
# '1월','12월','2025년' 같은 형태(토크나이저가 숫자/한글을 분리하므로 이 패턴은 주로 조합 확인용)
_NUM_WITH_SUFFIX = re.compile(r"^\d{1,4}[가-힣]$")


def _is_numeric_or_date_token(tok: str) -> bool:
    t = tok.strip()
    if not t:
        return True
    # 순수 숫자
    if _NUMERIC_FULL.match(t):
        return True
    # 요일/월 같은 단문 한글 (월/화/수…)
    if t in STOP_DAYS:
        return True
    # '1월' '12월' '2025년' 패턴 (토큰화에 따라 분해되더라도, 숫자/단일 한글이 각각 따로 필터링됨)
    if _NUM_WITH_SUFFIX.match(t):
        ch = t[-1]
        if ch in STOP_MONTH_SUFFIX:
            return True
    return False


def _is_stopword(tok_lc: str) -> bool:
    if tok_lc in WHITELIST:
        return False
    if tok_lc in STOP_EN:
        return True
    # 한국어는 소문자화 개념이 약하므로 두 세트 모두 확인
    if tok_lc in {s.lower() for s in STOP_KO}:
        return True
    return False


def _tokenize(text: str) -> List[str]:
    """
    한/영/숫자 연속 시퀀스를 토큰화한 후,
    - 순수 숫자/날짜류 제거
    - 시간표/행정 불용어 제거
    - 기술 화이트리스트는 보존
    """
    if not text:
        return []
    raw = _TOKEN_RE.findall(text)
    out: List[str] = []
    for t in raw:
        lc = t.lower()
        # 숫자/날짜/요일/월 등 제거
        if _is_numeric_or_date_token(t):
            continue
        # 일반 불용어 제거
        if _is_stopword(lc):
            continue
        # 한 글자 영문은 대부분 의미 빈약 — 단, 화이트리스트 'c'는 허용
        if (
            len(lc) == 1 and lc not in {"c"} and not ("\uac00" <= t <= "\ud7a3")
        ):  # 한글 한 글자는 유지
            continue
        out.append(lc)
    return out


# --------------------------
# 공개 클래스
# --------------------------
class TopicIndex:
    """
    간단 BM25 인덱스 래퍼
      - terms: 프롬프트 등에 쓸 수 있는 유니크 키워드 스트링 목록
      - scale: 스코어 squash 상수 (기본 6.0)
    """

    def __init__(
        self, docs_tokens: List[List[str]], raw_terms: List[str], scale: float = 6.0
    ) -> None:
        self.docs_tokens = docs_tokens
        self.terms: List[str] = sorted(set(raw_terms))
        self.scale = float(scale)
        self._bm25 = BM25Okapi(self.docs_tokens) if self.docs_tokens else None

    def topic_score(self, text: str) -> float:
        """
        입력 텍스트의 커리큘럼 근접도를 0~1 범위로 반환.
        - 문서별 BM25 점수 중 최대값을 1 - exp(-max/scale)로 squash.
        """
        if not text or not self._bm25:
            return 0.0
        q = _tokenize(text)
        if not q:
            return 0.0
        scores = self._bm25.get_scores(q)
        max_s = float(np.max(scores)) if len(scores) else 0.0
        return float(1.0 - math.exp(-max_s / self.scale))

    def delta_topic(self, before: str, after: str) -> float:
        return self.topic_score(after) - self.topic_score(before)


# --------------------------
# 빌더 (시간표 우선 → 일반 폴백)
# --------------------------
def build_topic_index(curriculum_xlsx_path: str) -> TopicIndex:
    """
    커리큘럼 XLSX에서 키워드/제목을 읽어 BM25 인덱스를 구축.
    - 시간표 형식 시트가 있으면 그 표만 사용
    - 없으면 일반 형식(키워드/과목/소단원) 파서로 시도
    """
    rows = _read_curriculum_rows_timetable_first(curriculum_xlsx_path)
    docs_tokens: List[List[str]] = []
    raw_terms: List[str] = []

    for rec in rows:
        kw_text = rec.get("keywords", "")
        base_text = (
            kw_text
            if kw_text.strip()
            else f"{rec.get('subject','')} {rec.get('unit','')}"
        )
        toks = _tokenize(base_text)
        if not toks:
            continue
        docs_tokens.append(toks)
        raw_terms.extend(toks)

    if not docs_tokens:
        # 완전 빈 경우라도 안전 인스턴스 반환
        return TopicIndex(docs_tokens=[], raw_terms=[])

    return TopicIndex(docs_tokens=docs_tokens, raw_terms=raw_terms, scale=6.0)


def build_curriculum_sentence_index(curriculum_xlsx_path: str) -> TopicIndex:
    """
    커리큘럼 XLSX에서 (교과목명, 세부 항목 요약, 세부내용)을 하나로 묶은 문장으로 BM25 인덱스를 구축.
    - 각 커리큘럼 항목을 하나의 완전한 문장으로 구성하여 더 정확한 매칭 제공
    - 시간표 형식 시트가 있으면 그 표만 사용
    - 없으면 일반 형식 파서로 시도
    """
    rows = _read_curriculum_rows_timetable_first(curriculum_xlsx_path)
    docs_tokens: List[List[str]] = []
    raw_terms: List[str] = []
    curriculum_sentences: List[str] = []

    for rec in rows:
        # 교과목명, 세부 항목 요약, 세부내용을 하나의 문장으로 구성
        subject = rec.get("subject", "").strip()
        unit = rec.get("unit", "").strip()
        detail = rec.get("detail", "").strip()
        
        # 키워드 필드에서 세부내용 추출 시도
        if not detail:
            keywords = rec.get("keywords", "").strip()
            # 키워드에서 교과목명과 세부 항목 요약을 제외한 나머지를 세부내용으로 간주
            if keywords and subject and unit:
                # 키워드에서 subject와 unit을 제거한 나머지를 detail로 사용
                remaining = keywords.replace(subject, "").replace(unit, "").strip()
                if remaining:
                    detail = remaining
        
        # 문장 구성: "교과목명의 세부 항목 요약에 대해 세부내용을 학습합니다"
        if subject and unit and detail:
            sentence = f"{subject}의 {unit}에 대해 {detail}을 학습합니다"
        elif subject and unit:
            sentence = f"{subject}의 {unit}을 학습합니다"
        elif subject:
            sentence = f"{subject}을 학습합니다"
        else:
            # 기존 방식으로 폴백
            sentence = rec.get("keywords", "").strip()
            if not sentence:
                continue
        
        if not sentence.strip():
            continue
            
        # 문장을 토큰화
        toks = _tokenize(sentence)
        if not toks:
            continue
            
        docs_tokens.append(toks)
        raw_terms.extend(toks)
        curriculum_sentences.append(sentence)

    if not docs_tokens:
        # 완전 빈 경우라도 안전 인스턴스 반환
        return TopicIndex(docs_tokens=[], raw_terms=[])

    # TopicIndex에 커리큘럼 문장 정보 추가
    topic_index = TopicIndex(docs_tokens=docs_tokens, raw_terms=raw_terms, scale=6.0)
    topic_index.curriculum_sentences = curriculum_sentences
    return topic_index


# --------------------------
# 시간표 전용 파서
# --------------------------
_EXPECT_HEADERS = [
    "주차",
    "일자",
    "요일",
    "구분",
    "교과목명",
    "세부 항목 요약",
    "세부내용",
]


def _read_curriculum_rows_timetable_first(xlsx_path: str) -> List[Dict[str, str]]:
    wb = load_workbook(xlsx_path, data_only=True)
    sheetnames = wb.sheetnames

    # 1) 시간표 후보 시트 우선순위: 이름에 '시간표' 또는 '(CLD4)'가 포함
    timetable_sheets = [n for n in sheetnames if "시간표" in n]
    # 폴백: 모든 시트 순회
    search_order = timetable_sheets + [
        n for n in sheetnames if n not in timetable_sheets
    ]

    for name in search_order:
        ws = wb[name]
        found = _try_parse_timetable_sheet(ws)
        if found:
            return found

    # 전부 실패 → 일반 파서
    return _read_curriculum_rows_generic(wb)


def _cell_str(v: Any) -> str:
    """셀 값을 문자열로 안정 변환 (날짜는 YYYY-MM-DD로)."""
    if v is None:
        return ""
    if isinstance(v, datetime):
        return v.strftime("%Y-%m-%d")
    return str(v).strip()


def _row_to_list(ws, row_idx: int, max_cols: int = 30) -> List[str]:
    """row_idx(1-base)의 셀을 왼쪽부터 문자열로 수집."""
    row = ws[row_idx]
    vals: List[str] = []
    for i, cell in enumerate(row[:max_cols]):
        vals.append(_cell_str(cell.value))
    return vals


def _find_header_row(
    ws, max_scan_rows: int = 60
) -> Optional[Tuple[int, Dict[str, int]]]:
    """
    상단 일부 행을 스캔하여 헤더 행과 '왼쪽 표'의 컬럼 인덱스를 찾는다.
    반환: (header_row_idx, colmap)
    colmap keys: "주차","일자","요일","구분","교과목명","세부 항목 요약","세부내용" (있는 것만)
    """
    for r in range(1, max_scan_rows + 1):
        vals = _row_to_list(ws, r, max_cols=30)
        hits = {h: i for i, v in enumerate(vals) for h in _EXPECT_HEADERS if v == h}
        if len(hits) >= 3 and ("교과목명" in hits):
            return r, hits
    return None


def _try_parse_timetable_sheet(ws) -> Optional[List[Dict[str, str]]]:
    """
    '표로 정리된 시간표'를 파싱하여 rows[{"keywords","subject","unit"}] 목록을 만든다.
    실패하면 None.
    """
    hdr = _find_header_row(ws)
    if hdr is None:
        return None
    header_row, colmap = hdr

    idx_week = colmap.get("주차")
    idx_date = colmap.get("일자")
    idx_wday = colmap.get("요일")
    idx_part = colmap.get("구분")
    idx_course = colmap.get("교과목명")
    idx_sum = colmap.get("세부 항목 요약")
    idx_detail = colmap.get("세부내용")

    if idx_course is None:
        return None

    rows_out: List[Dict[str, str]] = []
    max_rows = ws.max_row

    for r in range(header_row + 1, max_rows + 1):
        vals = _row_to_list(ws, r, max_cols=30)
        course = (
            vals[idx_course]
            if idx_course is not None and idx_course < len(vals)
            else ""
        )
        summary = vals[idx_sum] if (idx_sum is not None and idx_sum < len(vals)) else ""
        detail = (
            vals[idx_detail]
            if (idx_detail is not None and idx_detail < len(vals))
            else ""
        )
        part = vals[idx_part] if (idx_part is not None and idx_part < len(vals)) else ""

        # 빈 행/합계 행 배제
        if not (course or summary or detail or part):
            continue
        if course and ("시수 합계" in course or course.strip() == "합계"):
            continue

        fields = [course, summary, detail]
        # 보조 신호(가벼운 힌트) — 토큰화 단계에서 불용어로 자연스럽게 걸러짐
        for idx in (idx_part, idx_wday, idx_week, idx_date):
            if idx is not None and idx < len(vals) and vals[idx]:
                fields.append(vals[idx])

        keywords = " ".join([f for f in fields if f]).strip()
        if not keywords:
            continue

        rows_out.append(
            {"keywords": keywords, "subject": course or "", "unit": summary or "", "detail": detail or ""}
        )

    return rows_out if rows_out else None


# --------------------------
# 일반(폴백) 파서
# --------------------------
def _read_curriculum_rows_generic(wb) -> List[Dict[str, str]]:
    rows: List[Dict[str, str]] = []
    for ws in wb.worksheets:
        collected = _parse_generic_sheet(ws)
        rows.extend(collected)
    return rows


def _parse_generic_sheet(ws) -> List[Dict[str, str]]:
    header_row_idx = None
    header_lc: List[str] = []
    max_scan = min(ws.max_row, 50)

    for r in range(1, max_scan + 1):
        vals = [_cell_str(c.value).lower() for c in ws[r][:30]]
        if not any(vals):
            continue
        keys = ("키워드", "keywords", "과목", "subject", "소단원", "unit", "section")
        if sum(1 for v in vals if v in keys) >= 1:
            header_row_idx = r
            header_lc = vals
            break

    if header_row_idx is None:
        return []

    def _find_col(names: List[str]) -> Optional[int]:
        for n in names:
            nlc = n.lower()
            if nlc in header_lc:
                return header_lc.index(nlc)
        return None

    idx_keywords = _find_col(["키워드", "keywords"])
    idx_subject = _find_col(["과목", "subject"])
    idx_unit = _find_col(["소단원", "unit", "section"])

    out: List[Dict[str, str]] = []
    for r in range(header_row_idx + 1, ws.max_row + 1):
        vals = [_cell_str(c.value) for c in ws[r][:30]]
        kw = (
            vals[idx_keywords]
            if (idx_keywords is not None and idx_keywords < len(vals))
            else ""
        )
        subj = (
            vals[idx_subject]
            if (idx_subject is not None and idx_subject < len(vals))
            else ""
        )
        unit = vals[idx_unit] if (idx_unit is not None and idx_unit < len(vals)) else ""
        if not (kw or subj or unit):
            continue
        keywords = kw if kw.strip() else f"{subj} {unit}".strip()
        if keywords:
            out.append({"keywords": keywords, "subject": subj, "unit": unit})
    return out
